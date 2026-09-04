import os
import re
import json
import time
import requests
import openpyxl
from io import BytesIO
from urllib.parse import quote
from datetime import datetime, date

# ================= КОНФИГУРАЦИЯ =================
REPO_OWNER = "colderuopen-art"
REPO_NAME = "raspisanie"
FILE_PATH = "Расписание.xlsx"
TARGET_GROUP = "183р"
STATE_FILE = "state.json"

TG_BOT_TOKEN = os.getenv("TG_BOT_TOKEN")
TG_CHAT_ID = os.getenv("TG_CHAT_ID")

# Сетка звонков
BELL_SCHEDULE = {
    "ПОНЕДЕЛЬНИК": {
        "0 пара": "08:30 – 09:20",
        "1 пара": "09:20 – 10:40",
        "2 пара": "10:50 – 12:50",
        "3 пара": "13:00 – 14:20",
        "4 пара": "14:30 – 15:50",
        "5 пара": "16:00 – 17:20",
    },
    "DEFAULT": {  # Вторник – Пятница
        "1 пара": "08:30 – 09:50",
        "2 пара": "10:00 – 12:00",
        "3 пара": "12:10 – 13:30",
        "4 пара": "13:40 – 15:00",
        "5 пара": "15:10 – 16:30",
        "6 пара": "16:40 – 18:00",
    },
    "СУББОТА": {
        "1 пара": "08:30 – 09:50",
        "2 пара": "10:10 – 11:30",
        "3 пара": "11:50 – 13:10",
        "4 пара": "13:20 – 14:40",
        "5 пара": "14:50 – 16:10",
        "6 пара": "16:20 – 17:40",
    }
}

# Обеды для группы 183Р (литера "Р")
LUNCH_SCHEDULE = {
    "ПОНЕДЕЛЬНИК": "🥪 <i>Обед: 11:30 – 12:10</i>",
    "DEFAULT": "🥪 <i>Обед: 11:20 – 12:00</i>"
}

NUM_EMOJI = {
    "0": "0️⃣", "1": "1️⃣", "2": "2️⃣",
    "3": "3️⃣", "4": "4️⃣", "5": "5️⃣", "6": "6️⃣"
}

VALID_DAYS = ["ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА"]


def get_latest_commit_sha():
    """Проверка коммита через GitHub API"""
    url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/commits?path={quote(FILE_PATH)}&page=1&per_page=1"
    try:
        res = requests.get(url, timeout=15)
        if res.status_code == 200:
            data = res.json()
            if data:
                return data[0]["sha"]
    except Exception as e:
        print(f"Ошибка проверки коммита: {e}")
    return None


def download_excel():
    """Скачивание файла Excel в память"""
    url = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/main/{quote(FILE_PATH)}"
    res = requests.get(url, timeout=30)
    if res.status_code != 200:
        url = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/master/{quote(FILE_PATH)}"
        res = requests.get(url, timeout=30)
    res.raise_for_status()
    return BytesIO(res.content)


def extract_date_from_sheet(sheet):
    """Универсальное извлечение даты из первых трех строк вкладки"""
    for row in sheet.iter_rows(min_row=1, max_row=3, values_only=False):
        for cell in row:
            val = cell.value
            if not val:
                continue

            # 1. Если ячейка имеет встроенный формат даты Excel
            if isinstance(val, (datetime, date)):
                return val.strftime("%d.%m.%Y")

            val_str = str(val).strip()

            # 2. Поиск формата 3.9.2026 или 03.09.2026
            match_dot = re.search(r"\b\d{1,2}\.\d{1,2}\.\d{2,4}\b", val_str)
            if match_dot:
                return match_dot.group(0)

            # 3. Поиск формата 2026-09-03
            match_iso = re.search(r"\b\d{4}-\d{1,2}-\d{1,2}\b", val_str)
            if match_iso:
                try:
                    dt = datetime.strptime(match_iso.group(0), "%Y-%m-%d")
                    return dt.strftime("%d.%m.%Y")
                except Exception:
                    return match_iso.group(0)

    return ""


def parse_schedule(file_bytes):
    """Сбор расписания для группы 183Р"""
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    schedule_data = {}

    for sheet_name in wb.sheetnames:
        clean_day = sheet_name.strip().upper()
        if clean_day not in VALID_DAYS:
            continue

        sheet = wb[sheet_name]

        # 1. Надежный поиск даты
        date_str = extract_date_from_sheet(sheet)

        # 2. Поиск номеров пар во 2-й строке
        pair_columns = {}
        for col_idx, cell in enumerate(sheet[2], start=1):
            val = str(cell.value or "").strip().lower()
            if "пара" in val:
                pair_columns[col_idx] = cell.value.strip()

        # 3. Поиск строки группы 183р
        group_row_idx = None
        for row_idx in range(3, sheet.max_row + 1):
            cell_val = str(sheet.cell(row=row_idx, column=1).value or "").strip().lower()
            if cell_val == TARGET_GROUP or cell_val.startswith(TARGET_GROUP):
                group_row_idx = row_idx
                break

        if not group_row_idx:
            continue

        # 4. Сбор только заполненных пар
        day_pairs = {}
        for col_idx, pair_name in pair_columns.items():
            cell_val = sheet.cell(row=group_row_idx, column=col_idx).value
            if cell_val:
                clean_lesson = str(cell_val).strip()
                if clean_lesson and clean_lesson.lower() != "none":
                    day_pairs[pair_name] = clean_lesson

        schedule_data[clean_day] = {
            "date": date_str,
            "pairs": day_pairs
        }

    return schedule_data


def build_schedule_block(day_name, day_info):
    """Формирует список пар со временем (только фактические пары)"""
    if day_name == "ПОНЕДЕЛЬНИК":
        bells = BELL_SCHEDULE["ПОНЕДЕЛЬНИК"]
        lunch = LUNCH_SCHEDULE["ПОНЕДЕЛЬНИК"]
    elif day_name == "СУББОТА":
        bells = BELL_SCHEDULE["СУББОТА"]
        lunch = None
    else:
        bells = BELL_SCHEDULE["DEFAULT"]
        lunch = LUNCH_SCHEDULE["DEFAULT"]

    pairs = day_info.get("pairs", {})
    if not pairs:
        return "<i>На этот день пар нет</i>"

    def sort_key(item):
        num_part = item[0].split()[0]
        return int(num_part) if num_part.isdigit() else 99

    sorted_pairs = sorted(pairs.items(), key=sort_key)

    lines = []
    for pair_name, lesson in sorted_pairs:
        pair_num = pair_name.split()[0]
        emoji = NUM_EMOJI.get(pair_num, "🔹")
        time_range = bells.get(pair_name, "")
        time_part = f"<b>{time_range}</b> | " if time_range else ""

        lines.append(f"{emoji} {time_part}{lesson}")

        if pair_name == "2 пара" and lunch:
            lines.append(f"    ↳ {lunch}")

    return "\n".join(lines)


def format_new_date_message(day_name, day_info):
    """Сообщение о новом расписании"""
    date_str = day_info.get("date", "")
    header_date = f"{day_name}, {date_str}".strip(", ")
    schedule_text = build_schedule_block(day_name, day_info)
    now_time = datetime.now().strftime("%d.%m.%Y в %H:%M")

    return (
        f"📅 <b>Опубликовано расписание на дату!</b>\n\n"
        f"👥 <b>Группа:</b> {TARGET_GROUP.upper()}\n"
        f"📆 <b>{header_date}</b>\n\n"
        f"───────────────────\n"
        f"📋 <b>АКТУАЛЬНОЕ РАСПИСАНИЕ:</b>\n"
        f"{schedule_text}\n\n"
        f"<i>(Обновлено: {now_time})</i>"
    )


def format_replacement_message(day_name, day_info, changes):
    """Сообщение о замене"""
    date_str = day_info.get("date", "")
    header_date = f"{day_name}, {date_str}".strip(", ")

    diff_lines = []
    for pair_name, (old_val, new_val) in sorted(changes.items()):
        if old_val and new_val:
            diff_lines.append(f"• <b>{pair_name}:</b> <s>{old_val}</s> ➔ <b>{new_val}</b>")
        elif not old_val and new_val:
            diff_lines.append(f"• <b>{pair_name}:</b> <i>добавлена</i> ➔ <b>{new_val}</b>")
        elif old_val and not new_val:
            diff_lines.append(f"• <b>{pair_name}:</b> <s>{old_val}</s> ➔ <i>отменена</i>")

    diff_text = "\n".join(diff_lines)
    schedule_text = build_schedule_block(day_name, day_info)
    now_time = datetime.now().strftime("%d.%m.%Y в %H:%M")

    return (
        f"🔔 <b>Внимание! Изменение в расписании</b>\n\n"
        f"👥 <b>Группа:</b> {TARGET_GROUP.upper()}\n"
        f"📅 <b>{header_date}</b>\n\n"
        f"⚠️ <b>ЧТО ИЗМЕНИЛОСЬ:</b>\n"
        f"{diff_text}\n\n"
        f"───────────────────\n"
        f"📋 <b>АКТУАЛЬНОЕ РАСПИСАНИЕ НА ДЕНЬ:</b>\n"
        f"{schedule_text}\n\n"
        f"<i>(Обновлено: {now_time})</i>"
    )


def send_telegram_notification(text):
    """Отправка сообщения в чат Telegram"""
    if not TG_BOT_TOKEN or not TG_CHAT_ID:
        print("Ошибка: TG_BOT_TOKEN или TG_CHAT_ID не заданы!")
        return
    url = f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": TG_CHAT_ID,
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": True
    }
    res = requests.post(url, json=payload, timeout=10)
    res.raise_for_status()


def main():
    state = {}
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                state = json.load(f)
        except Exception:
            state = {}

    last_sha = state.get("last_commit_sha")
    old_schedule = state.get("schedule", {})

    current_sha = get_latest_commit_sha()
    if current_sha and current_sha == last_sha:
        print("Файл не менялся. Завершение работы.")
        return

    print("Скачиваем актуальный Excel...")
    excel_bytes = download_excel()
    new_schedule = parse_schedule(excel_bytes)

    # Если это первый запуск или старые даты были пустыми — обновляем снимок
    has_empty_old_dates = any(not info.get("date") for info in old_schedule.values()) if old_schedule else True

    if not old_schedule or has_empty_old_dates:
        print("Инициализация/исправление базы дат в state.json...")
        state["last_commit_sha"] = current_sha
        state["schedule"] = new_schedule
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        return

    has_notified = False

    for day, new_info in new_schedule.items():
        old_info = old_schedule.get(day, {"pairs": {}, "date": ""})
        old_pairs = old_info.get("pairs", {})
        new_pairs = new_info.get("pairs", {})

        old_date = old_info.get("date", "").strip()
        new_date = new_info.get("date", "").strip()

        # СИТУАЦИЯ 1: Сменилась дата
        if new_date and old_date and new_date != old_date:
            if new_pairs:
                print(f"Новая дата для {day}: {new_date}")
                msg = format_new_date_message(day, new_info)
                send_telegram_notification(msg)
                has_notified = True
                time.sleep(1)
            continue

        # СИТУАЦИЯ 2: Точечные замены
        all_pair_names = set(old_pairs.keys()).union(new_pairs.keys())
        day_changes = {}

        for p_name in all_pair_names:
            old_val = old_pairs.get(p_name)
            new_val = new_pairs.get(p_name)
            if old_val != new_val:
                day_changes[p_name] = (old_val, new_val)

        if day_changes:
            print(f"Замены на {day} для 183р")
            msg = format_replacement_message(day, new_info, day_changes)
            send_telegram_notification(msg)
            has_notified = True
            time.sleep(1)

    if not has_notified:
        print("Файл обновился, но группу 183Р изменения не затронули.")

    state["last_commit_sha"] = current_sha
    state["schedule"] = new_schedule
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
